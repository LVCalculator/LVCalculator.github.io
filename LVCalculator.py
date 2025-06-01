from spire.xls.common import *
from spire.xls import *
from spire.pdf.common import *
from spire.pdf import *
import spire.xls
# import spire.pdf
import pandas as pd
import numpy as np
import os
import openpyxl
# from openpyxl import load_workbook

thin_border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'), 
                     right=openpyxl.styles.borders.Side(style='thin'), 
                     top=openpyxl.styles.borders.Side(style='thin'), 
                     bottom=openpyxl.styles.borders.Side(style='thin'))

lv_raw = "C:\\Users\\mrs69\\Desktop\\Programs\\LVCalculator\\ErdbauSVCalc\\LV\\Preisliste HHR Glasfaser OXG_Erdbausv_v3.pdf"
lv_path = "C:\\Users\\mrs69\\Desktop\\Programs\\LVCalculator\\ErdbauSVCalc\\LV"
position_names = ["TB-", "ME-"]

def pdf2excel(file_name, excel_path):
    pdf = PdfDocument()
    # Load a PDF document
    pdf.LoadFromFile(file_name)
    convertOptions = XlsxLineLayoutOptions(False, True, False, True, False)

    # Set the conversion options
    pdf.ConvertOptions.SetPdfToXlsxOptions(convertOptions)

    pdf.SaveToFile(excel_path, FileFormat.XLSX)
    pdf.Close()


def get_lv(path_to_lv):
    # if lv_raw[-4:] == ".pdf":
    #     pdf2excel(lv_raw, lv_path)
    
    df = pd.read_excel(path_to_lv).dropna(axis=1, how='all')
    lv_df = pd.DataFrame(columns=["Code", "Description", "Quantity", "Price"])
    for index, row in df.iterrows():
        clean_row = row.dropna()
        if len(clean_row) < 3: continue
        if clean_row.iloc[0].strip()[:3] not in position_names:
            continue
        code, desc = clean_row.iloc[0].strip(), clean_row.iloc[1].strip()
        if len(clean_row) == 4 and "€" in clean_row.iloc[3]:
            quant, price = clean_row.iloc[2].strip(), float(clean_row.iloc[3].strip()[:-2].replace(',','.'))
        elif "€" in clean_row.iloc[2]:
            quant, price = np.nan, float(clean_row.iloc[2].strip()[:-2])
        else:
            quant, price = clean_row.iloc[2].strip(), np.nan
        lv_df.loc[-1] = [code, desc, quant, price]
        lv_df.index = lv_df.index + 1
        lv_df = lv_df.sort_index()
         
    return lv_df

def get_aufmass(file_name):
    aufmass_path = 'aufmass_excel.xlsx'
    pdf2excel(file_name, 'aufmass_excel.xlsx')
    df = pd.read_excel(aufmass_path).dropna(axis=1, how='all')
    flip = False
    for index, row in df.iterrows():
        clean_row = row.dropna()
        if len(clean_row) < 2:
            continue
        some_ctr = 0
        for i in range(len(clean_row)):
            if clean_row.iloc[i].strip()[:3] in position_names:
                some_ctr += 1
        if some_ctr > 3: 
            flip = True
            break

    if flip:
        df = df.transpose()

    aufmass_df = pd.DataFrame(columns=["Code", "Ammount"])
    for index, row in df.iterrows():
        clean_row = row.dropna()
        if len(clean_row) < 2:
            continue
        if clean_row.iloc[0].strip()[:3] not in position_names:
            continue
        code = clean_row.iloc[0].strip().replace("\n","").replace(" ","")
        amt = 0
        nr_ctr = 0
        for i in range(1, len(clean_row)):
            f = clean_row.iloc[i].strip().replace(",",".")
            ctr = 0
            d = "0"
            while ctr < len(f) and f[ctr] in '0123456789.':
                d = d + f[ctr]
                ctr += 1
            if float(d) > 0: 
                nr_ctr += 1
            amt += float(d)
            if f[:3] in position_names:
                break
        
        if code[:6] == "TB-026":
            amt = nr_ctr

        aufmass_df.loc[-1] = [code, amt]
        aufmass_df.index = aufmass_df.index + 1
        aufmass_df = aufmass_df.sort_index()
    return aufmass_df


def run_processor(lv_path, aufmass_path, output_path, file_name):
    df_aufmass = get_aufmass(aufmass_path)
    df_lv = get_lv(os.path.join(lv_path, 'current_LV.xlsx'))
    d = []
    for i in range(len(df_aufmass)):
        code = df_aufmass['Code'].iloc[i]
        best_code = ""
        for j in range(min(10, len(code)+1)):
            if code[:j] in set(df_lv["Code"]):
                best_code = code[:j]
        if best_code == "":
            print("No such code:", code)
        
        d.append(best_code)
    df_aufmass['Code'] = d
    df_aufmass = df_aufmass.groupby("Code").sum()
    df_aufmass = df_aufmass[df_aufmass["Ammount"] > 0]
    df_aufmass = df_aufmass.merge(df_lv, how='inner', on='Code')
    df_aufmass["Value"] = df_aufmass["Ammount"] * df_aufmass["Price"]
    df_aufmass = df_aufmass.sort_values(by=['Code'])
    
    wb = openpyxl.load_workbook(os.path.join(lv_path, 'RechnungVorlageESV.xlsx'))
    ws = wb.active
    ws['G26'] = f"{df_aufmass['Value'].sum():.2f}€"
    ws.insert_rows(idx=22, amount=len(df_aufmass))
    for index, row in df_aufmass.iterrows():
        ctr = str(22 + index)
        ws['A' + ctr] = row["Code"]
        ws['B' + ctr] = row["Description"]
        ws['B' + ctr].alignment = openpyxl.styles.Alignment(wrap_text=True)
        ws['D' + ctr] = f'{row["Ammount"]:.2f}'
        ws['D' + ctr].alignment = openpyxl.styles.Alignment(horizontal='right')
        ws['E' + ctr] = f'{row["Price"]:.2f}€'
        ws['E' + ctr].alignment = openpyxl.styles.Alignment(horizontal='right')
        ws['G' + ctr] = f'{row["Value"]:.2f}€'
        ws['G' + ctr].alignment = openpyxl.styles.Alignment(horizontal='right')
        
        ws['A' + ctr].border = thin_border
        ws['B' + ctr].border = thin_border
        ws['C' + ctr].border = thin_border
        ws['D' + ctr].border = thin_border
        ws['E' + ctr].border = thin_border
        ws['F' + ctr].border = thin_border
        ws['G' + ctr].border = thin_border

    # File name: Street Number Contractor-Contracted-CalendarWeek-Project-Number
    # File name: Hausanschluss Contractor-Contracted-CalendarWeek-Project-Number
    
    *street, receipt_code = file_name.split()
    ag, an, date, project, nr = receipt_code.split('-')
    ws['B18'] = f'Trassenherstellung für {ag}, {" ".join(street)}'
    ws['B19'] = f'{project}, {date}, {receipt_code}'
        

    output_file = os.path.join(output_path, file_name[:-4] + '.xlsx')
    wb.save(output_file)
    
    return output_file






# def run_processor(lv_path, upload_path, processed_path):

# get_prices(lv_path, upload_path, processed_path)


# get_prices(lv_path, "C:\\Users\\mrs69\\Desktop\\Programs\\LVCalculator\\ErdbauSVCalc\\Input\\ErdbauSVAufmassKW18-19.pdf", "C:\\Users\\mrs69\\Desktop\\Programs\\LVCalculator\\ErdbauSVCalc\\processed\\processed.xlsx")


# get_prices(lv_path, "C:\\Users\\mrs69\\Desktop\\Programs\\LVCalculator\\ErdbauSVCalc\\Input\\ErdbauSVHAListe5_15_25_signed.pdf")

# get_aufmass("C:\\Users\\mrs69\\Desktop\\Programs\\LVCalculator\\ErdbauSVCalc\\Input\\ErdbauSVAufmassKW18-19.pdf")

# get_aufmass("C:\\Users\\mrs69\\Desktop\\Programs\\LVCalculator\\ErdbauSVCalc\\Input\\ErdbauSVHAListe5_15_25_signed.pdf")


